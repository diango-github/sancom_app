from django import forms
from account.models import File
from django.contrib.auth.models import User
import openpyxl
from config.settings import BASE_DIR

class Excel_link:

    def __init__(self, filename, sheet):
        self.fname = filename
        self.sheet = sheet
        self.verb_dic={}
        self.verb_list=[]
        self.listbox=[]
        self.excel_data=[]
    
    def getlist(self):
                #wb=openpyxl.load_workbook("./static/" + self.fname)
        wb=openpyxl.load_workbook(BASE_DIR + '/static/' + self.fname)
        sh=wb.get_sheet_by_name(self.sheet)
        for row in range(2,sh.max_row +1 ):
            item=sh["B"+str(row)].value
            category=sh["C"+str(row)].value
            japanese=sh["D"+str(row)].value
            english=sh["E"+str(row)].value
            esound=sh["F"+str(row)].value
            eword1=sh["G"+str(row)].value
            eword2=sh["H"+str(row)].value
            eword3=sh["I"+str(row)].value
            chinese=sh["J"+str(row)].value
            csound=sh["K"+str(row)].value
            cword1=sh["L"+str(row)].value
            cword2=sh["M"+str(row)].value
            cword3=sh["N"+str(row)].value
            self.verb_dic.setdefault(item,{"category":category, "japanese":japanese, "english":english, "esound":esound,"eword1":eword1, "eword2":eword2, "eword3":eword3, "chinese":chinese,"csound":csound,"cword1":cword1,"cword2":cword2,"cword3":cword3})
        self.verb_list=list(self.verb_dic.keys())
        for i in range(len(self.verb_list)):
            self.listbox.append((self.verb_list[i], self.verb_list[i]))
        self.excel_data = [self.verb_dic, self.verb_list, self.listbox]
        return self.excel_data

class Lan_appForm1(forms.Form):

    def __init__(self, filename, sheet, *args, **kwargs):
        super(Lan_appForm1, self).__init__(*args, **kwargs)
        self.fname = filename
        excelfile = self.fname
        self.sheet = sheet
        excelsheet = self.sheet
        excel = Excel_link(excelfile, excelsheet)
        excel_list = excel.getlist()
        self.fields['choice1'] = forms.ChoiceField(label='', choices=excel_list[2], required=True, widget=forms.Select(attrs={'id':'choice1','size':4}))

class En_form1(forms.Form):

    def __init__(self, en_list1=[], *args, **kwargs):
        super(En_form1, self).__init__(*args, **kwargs)
        self.fields['en_choice1'] = forms.ChoiceField(label='英語', choices=en_list1, widget=forms.Select(attrs={'id':'en_choice1','size':1}))

class Ch_form1(forms.Form):

    def __init__(self, ch_list1=[], *args, **kwargs):
        super(Ch_form1, self).__init__(*args, **kwargs)
        self.fields['ch_choice1'] = forms.ChoiceField(label='中国語', choices=ch_list1, widget=forms.Select(attrs={'id':'ch_choice1','size':1}))

class En_form2(forms.Form):

    def __init__(self, en_list2=[], *args, **kwargs):
        super(En_form2, self).__init__(*args, **kwargs)
        self.fields['en_choice2'] = forms.ChoiceField(label='英語', choices=en_list2, widget=forms.Select(attrs={'id':'en_choice2','size':1}))

class Ch_form2(forms.Form):

    def __init__(self, ch_list2=[], *args, **kwargs):
        super(Ch_form2, self).__init__(*args, **kwargs)
        self.fields['ch_choice2'] = forms.ChoiceField(label='中国語', choices=ch_list2, widget=forms.Select(attrs={'id':'ch_choice2','size':1}))

class En_form3(forms.Form):

    def __init__(self, en_list3=[], *args, **kwargs):
        super(En_form3, self).__init__(*args, **kwargs)
        self.fields['en_choice3'] = forms.ChoiceField(label='英語', choices=en_list3, widget=forms.Select(attrs={'id':'en_choice3','size':1}))

class Ch_form3(forms.Form):

    def __init__(self, ch_list3=[], *args, **kwargs):
        super(Ch_form3, self).__init__(*args, **kwargs)
        self.fields['ch_choice3'] = forms.ChoiceField(label='中国語', choices=ch_list3, widget=forms.Select(attrs={'id':'ch_choice3','size':1}))

class En_form4(forms.Form):

    def __init__(self, en_list4=[], *args, **kwargs):
        super(En_form4, self).__init__(*args, **kwargs)
        self.fields['en_choice4'] = forms.ChoiceField(label='英語', choices=en_list4, widget=forms.Select(attrs={'id':'en_choice4','size':1}))

class Ch_form4(forms.Form):

    def __init__(self, ch_list4=[], *args, **kwargs):
        super(Ch_form4, self).__init__(*args, **kwargs)
        self.fields['ch_choice4'] = forms.ChoiceField(label='中国語', choices=ch_list4, widget=forms.Select(attrs={'id':'ch_choice4','size':1}))

class DeepenRadio(forms.Form):
    data = (
        (1, '前回結果の悪い順'),
        (2, '訓練回数の少ない順'),
        (3, '経過日数の多い順')
    )
    choice = forms.ChoiceField(label='優先順位並び替え', choices=data, initial='3', widget=forms.RadioSelect())

class DeepenRadio2(forms.Form):
    data = (
        (5, 'レベル５：ネイティブの速さで言えた'),
        (4, 'レベル４：自分なりにすらすら言えた'),
        (3, 'レベル３：たどたどしく言えた'),
        (2, 'レベル２：一部を間違えた'),
        (1, 'レベル１：思い出せなかった')
    )
    choice = forms.ChoiceField(label='', initial='1', choices=data, widget=forms.RadioSelect())

class FinishRadio(forms.Form):
    data = (
        (1, '１週間以上１ヶ月未満の経過'),
        (2, '１ヶ月以上３カ月未満の経過'),
        (3, '３ヶ月以上の経過')
    )
    choice = forms.ChoiceField(label='確認したい文例の抽出条件を指定します', initial='3', choices=data, widget=forms.RadioSelect())

class FinishRadio2(forms.Form):
    data = (
        (3, '覚えていた'),
        (2, '苦労して想い出した'),
        (1, '忘れていた')
    )
    choice = forms.ChoiceField(label='', initial='3', choices=data, widget=forms.RadioSelect())